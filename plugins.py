from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import config
import os
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

def read_data(filename=config.Input_file):
    """
    Reads data from an Excel file and returns parameters and data arrays.

    Parameters:
        filename (str): Name of the Excel file.

    Returns:
        tuple: CF, U, H, V, CP, CH, G, CT, D, pr, demand
    """
    IS = config.IS
    AS = config.AS
    NS = config.NS
    Food = config.Food
    Medicine = config.Medicine

    scenario_sheet_name = f'scenario_{IS}'
    probability_column_name = f'w{NS // 100}'

    scenario_data = pd.read_excel(filename, scenario_sheet_name)
    main_data = pd.read_excel(filename, 'main_data', header=None)

    # probability
    pr = scenario_data[probability_column_name].to_numpy()

    # demand
    demand = scenario_data.iloc[:NS, 1:IS + 1].to_numpy()
    demand = np.rint(demand * 0.03)
    Dr = np.empty((NS, IS, AS))
    Dr[:, :, 0] = scenario_data.iloc[:NS, 1:IS + 1].to_numpy()
    Dr[:, :, 1] = np.rint((Dr[:, :, 0] * Food))
    Dr[:, :, 2] = np.rint((Dr[:, :, 0] * Medicine))

    # Fixed cost, Storage capacity, volume of items
    CF = main_data.loc[1, 1:3].to_numpy()
    U = main_data.loc[2, 1:3].to_numpy()

    # Distance
    H = main_data.loc[8:(7 + IS), 1:].to_numpy()

    # volume of items
    V = main_data.loc[1, 6:8].to_numpy()
    # Unit procurement price
    CP = main_data.loc[2, 6:8].to_numpy()
    # Unit transportation cost
    CT = main_data.loc[3, 6:8].to_numpy()
    # Unit holding cost
    CH = main_data.loc[4, 6:8].to_numpy()
    # Unit penalty cost
    G = main_data.loc[5, 6:8].to_numpy()

    D = np.zeros((NS, AS, IS))
    for j in range(IS):
        for a in range(AS):
            for s in range(NS):
                D[s, a, j] = np.rint(Dr[s, j, a] * 0.03)

    return CF, U, H, V, CP, CH, G, CT, D, pr, demand

def generate_script_name(data_preprocess_methods, cluster_methods, sample_methods):
    if data_preprocess_methods:
        script_name = f"{data_preprocess_methods}_{cluster_methods}_{sample_methods}"
    else:
        script_name = f"{cluster_methods}_{sample_methods}"
    return script_name

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None ,index=False ,header=False , **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing Excel file
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be a dictionary]

    Returns: None

    (Example)
    append_df_to_excel('output.xlsx', df)

    """
    # Load existing workbook or create a new one
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(filename)

    # Check if sheet exists, if not create it
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    # Get the active sheet
    sheet = wb[sheet_name]

    # Calculate starting row
    if startrow is None:
        startrow = sheet.max_row if sheet.max_row > 0 else 1

    # Convert the DataFrame to rows
    # rows = list(dataframe_to_rows(df, header=True))
    rows = list(dataframe_to_rows(df, index=index, header=header))

    # Write DataFrame rows to Excel sheet
    for row in rows:
        sheet.append(row)

    # Save the workbook
    wb.save(filename)

def save_detailed_results(script_name, Vx, Vy, elapsed_time, start_row=1, sheet_name='result'):
    """
    Save the detailed computation results to an Excel file.

    :param filename: str, name of the Excel file.
    :param script_name: str, name of the script/method used.
    :param Vx: list, location data.
    :param Vy: list, inventory data.
    :param elapsed_time: float, elapsed time of the computation.
    :param start_row: int, start row for the first DataFrame.
    :param sheet_name: str, sheet name to write the data into.
    """
    # 创建DataFrame来组织数据
    costs_df = pd.DataFrame([[script_name, elapsed_time]], columns=['Method', 'Costs'])
    location_df = pd.DataFrame(Vx)
    inventory_df = pd.DataFrame(Vy).T
    elapsed_time_df = pd.DataFrame([['Elapsed time', elapsed_time]])

    # Save costs data
    append_df_to_excel(config.Input_file, costs_df, sheet_name=sheet_name, index=False, header=False, startrow=start_row)

    # Calculate next start column for location data
    location_startcol = costs_df.shape[1] + 2  # Assuming 2 column gap for readability
    append_df_to_excel(config.Input_file, location_df, sheet_name=sheet_name, index=True, header=True, startrow=start_row, startcol=location_startcol)

    # Calculate next start column for inventory data
    inventory_startcol = location_startcol + location_df.shape[1] + 2  # Assuming 2 column gap for readability
    append_df_to_excel(config.Input_file, inventory_df, sheet_name=sheet_name, index=True, header=True, startrow=start_row, startcol=inventory_startcol)

    # Calculate start row for elapsed time data
    elapsed_time_startrow = start_row + max(costs_df.shape[0], inventory_df.shape[0], location_df.shape[0]) + 2  # Assuming 2 row gap for readability
    append_df_to_excel(config.Input_file, elapsed_time_df, sheet_name=sheet_name, index=True, header=True, startrow=elapsed_time_startrow)
    
    print("Detailed results have been saved to Excel.")

def plot_cluster_sampling(demand_transformed, cluster_labels, sample, save_directory, script_name, m):
    
    # 绘制聚类结果和分层采样的样本点
    plt.figure(figsize=(10, 8),dpi=300)
    for i, label in enumerate(np.unique(cluster_labels)):
        plt.scatter(
            demand_transformed[cluster_labels == label, 0],
            demand_transformed[cluster_labels == label, 1],
            label=f'Cluster {i}'
        )

    # 假设sample包含了所有采样点的索引
    for s in sample:
        plt.scatter(
            demand_transformed[s, 0],
            demand_transformed[s, 1],
            c='red', # 标记颜色
            edgecolor='k',
            marker='x', # 标记样式
            s=100 # 设置标记大小为100 point^
        )

    plt.title('Stratified Sampling of Clusters')
    plt.xlabel('PCA Feature 1')
    plt.ylabel('PCA Feature 2')
    plt.legend()

    # 检查保存目录是否存在，如果不存在则创建
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # 保存图片到指定目录
    plt.savefig(os.path.join(save_directory, f'{script_name}_sample_{m+1}.jpg'), format='jpg')
    # plt.close()  # 关闭图形界面，防止内存泄露

def plot_cluster_3d_sampling(demand_transformed, cluster_labels, sample, save_directory, script_name, m):
    # 创建一个新的图和一个三维轴
    fig = plt.figure(figsize=(10, 8), dpi=300)
    ax = fig.add_subplot(111, projection='3d')

    # 绘制聚类结果的散点图
    for i, label in enumerate(np.unique(cluster_labels)):
        ax.scatter(
            demand_transformed[cluster_labels == label, 0],  # x轴坐标
            demand_transformed[cluster_labels == label, 1],  # y轴坐标
            demand_transformed[cluster_labels == label, 2],  # z轴坐标
            label=f'Cluster {i}'
        )

    # 假设sample包含了所有采样点的索引
    for s in sample:
        ax.scatter(
            demand_transformed[s, 0],  # x轴坐标
            demand_transformed[s, 1],  # y轴坐标
            demand_transformed[s, 2],  # z轴坐标
            c='red',  # 标记颜色
            edgecolor='k',
            marker='x',  # 标记样式
            s=100  # 设置标记大小为100 point^
        )

    # 设置图表标题和坐标轴标签
    ax.set_title('Stratified Sampling of Clusters')
    ax.set_xlabel('TSNE Feature 1')
    ax.set_ylabel('TSNE Feature 2')
    ax.set_zlabel('TSNE Feature 3')

    # 设置视角
    ax.view_init(elev=20., azim=45)  # 设置斜视图

    # 显示图例
    ax.legend()

    # 检查保存目录是否存在，如果不存在则创建
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # 保存图片到指定目录
    # save_path = os.path.join(save_directory, f'{script_name}_sample_3D_{m+1}.jpg')
    # plt.savefig(save_path, format='jpg')
    plt.savefig(os.path.join(save_directory, f'{script_name}_sample_3D_{m+1}.jpg'), format='jpg')
    # plt.close()  # 关闭图形界面，防止内存泄露
    # print(f"Saved plot as {save_path}")

def plot_cluster(demand_transformed, cluster_labels, save_directory, script_name):
    
    plt.figure(figsize=(10, 8), dpi=300)
    # 绘制聚类结果
    for i, label in enumerate(np.unique(cluster_labels)):
        plt.scatter(
            demand_transformed[cluster_labels == label, 0],
            demand_transformed[cluster_labels == label, 1],
            label=f'Cluster {i}'
        )

    plt.title('Clustering Results')
    plt.xlabel('Feature 1')
    plt.ylabel('Feature 2')
    plt.legend()

    # 检查保存目录是否存在，如果不存在则创建
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # 保存图片到指定目录
    plt.savefig(os.path.join(save_directory, f'{script_name}_cluster.jpg'), format='jpg')
    # plt.close()  # 关闭图形界面，防止内存泄露

def plot_cluster_3d(demand_transformed, cluster_labels, save_directory, script_name):
    # 创建一个新的图和一个三维轴
    fig = plt.figure(figsize=(10, 8), dpi=300)
    ax = fig.add_subplot(111, projection='3d')

    # 绘制聚类结果的散点图
    for i, label in enumerate(np.unique(cluster_labels)):
        ax.scatter(
            demand_transformed[cluster_labels == label, 0],  # x轴坐标
            demand_transformed[cluster_labels == label, 1],  # y轴坐标
            demand_transformed[cluster_labels == label, 2],  # z轴坐标
            label=f'Cluster {i}'
        )

    # 设置图表标题和坐标轴标签
    ax.set_title('3D Clustering Results')
    ax.set_xlabel('Feature 1')
    ax.set_ylabel('Feature 2')
    ax.set_zlabel('Feature 3')

    # 设置视角
    ax.view_init(elev=20., azim=45)

    # 显示图例
    ax.legend()

    # 检查保存目录是否存在，如果不存在则创建
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # 保存图片到指定目录
    save_path = os.path.join(save_directory, f'{script_name}_cluster_3D.jpg')
    plt.savefig(save_path, format='jpg')
    # plt.close()  # 关闭图形界面，防止内存泄露

def generate_cluster_plots(demand_transformed, samples_info, cluster_labels):
    # Generate cluster plots for each sample
    for sample, script_name, m in samples_info:
        if config.DIM_REDUCTION_METHOD == '2d':
            plot_cluster_sampling(demand_transformed, cluster_labels, sample, config.Graphs_sample_save_directory, script_name, m)
        elif config.DIM_REDUCTION_METHOD == '3d':
            plot_cluster_3d_sampling(demand_transformed, cluster_labels, sample, config.Graphs_sample_save_directory, script_name, m)
        else:
            raise ValueError("Invalid plot_type: {}. Choose '2d' or '3d'.".format(config.DIM_REDUCTION_METHOD))

def calculate_gap(ff, MS, gurobi_opt):
    """
    Calculate the GAP percentage between the average solution of the sample groups
    and the provided global best result.

    :param ff: List or array of solutions from the sample groups.
    :param MS: The number of samples or sample groups.
    :param gurobi_opt: The known global best result from Gurobi or similar solver.
    :return: GAP percentage.
    """
    # Check for valid inputs
    if MS <= 0:
        raise ValueError("Number of samples (MS) must be greater than 0.")
    if gurobi_opt <= 0:
        raise ValueError("Global best result (gurobi_opt) must be greater than 0.")
    
    # Calculate average of the solutions
    ave_f = np.sum(ff) / MS

    # Calculate GAP
    gap_percentage = (gurobi_opt - ave_f) / gurobi_opt * 100  # GAP as a percentage

    return gap_percentage

def save_and_print_results(script_name, IS, NS, MS, SS_SAA, opt_f, elapsed_time, cluster_num = 0, gap = 0):
    """
    Save results to an Excel file and print them.

    :param script_name: str, name of the script/method used.
    :param IS: int, input size parameter.
    :param NS: int, number of simulations/iterations/etc.
    :param opt_f: float, the optimized function value.
    :param elapsed_time: float, time taken for the operation in seconds.
    """
    # 创建一个DataFrame来组织需要输出的数据
    result_df = pd.DataFrame([[script_name, IS, NS, MS, SS_SAA, float(opt_f), elapsed_time, cluster_num, gap]])

    # 将数据输出到Excel的特定列，只有一列
    append_df_to_excel(config.Output_file, result_df, sheet_name='results', index=False, header=False, startrow=0)

    # 打印结果
    print(f"Method: {script_name}")
    print(f"I: {IS}, S: {NS}, M: {MS}, N: {SS_SAA}, clustering_num: {cluster_num}")    
    print(f"Costs: {float(opt_f)}, gap: {gap}")
    print(f"Elapsed time: {elapsed_time} seconds.")