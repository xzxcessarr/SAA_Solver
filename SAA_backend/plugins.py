# -*- coding: utf-8 -*-
"""
SAA_Solver
==========

This module is part of the SAA_Solver project and contains the implementation of Plugins.

Author: cessarr
Date Created: 2024-04-20
License: MIT License

Description:
------------
此部分为其他组件模块，包括数据读取、保存、聚类与抽样图片生成等
"""

import pandas as pd
import numpy as np
import os
import matplotlib.pyplot as plt
import redis
import json
import pickle


"""
三个store_data_to_redis都可以将Excel中的数据存储到Redis数据库。记住store_data_to_redis与read_data_from_redis一一对应

本函数读取指定的Excel文件，提取关键数据，并将其存储到Redis数据库中，以供后续快速访问和处理。支持原始数据和特定格式化数据的存储。

:param filename: Excel文件的路径和名称。
:param IS: 场景数量，用于确定读取Excel中的哪个场景数据。
:param Raw_data_flag: 原始数据标志，如果为True，则读取原始格式的数据；否则读取常规格式的数据。
"""
# Redis存储接口1
# def store_data_to_redis(filename, IS, Raw_data_flag):
#     # 连接到Redis
#     r = redis.Redis(host='localhost', port=6379, db=0, decode_responses=True)

#     # 读取Excel数据
#     if Raw_data_flag:
#         scenario_sheet_name = f'scenario_{20 if IS <= 20 else 40}'
#     else:
#         scenario_sheet_name = f'scenario'
    
#     scenario_data = pd.read_excel(filename, scenario_sheet_name)
#     distance_data = pd.read_excel(filename, 'distance', header=None)
#     facility_cost_data = pd.read_excel(filename, 'facility_cost', header=None)
#     resource_cost_data = pd.read_excel(filename, 'resource_cost', header=None)

#     # 存储到Redis
#     r.set('scenario_data', scenario_data.to_json())
#     r.set('distance_data', distance_data.to_json())
#     r.set('facility_cost_data', facility_cost_data.to_json())
#     r.set('resource_cost_data', resource_cost_data.to_json())

#     print("Data stored in Redis successfully!")

# Redis存储接口2
def store_data_to_redis(filename, IS, Raw_data_flag):
    # 连接到Redis
    r = redis.Redis(host='localhost', port=6379, db=0, decode_responses=True)

    # 读取Excel数据
    if Raw_data_flag:
        scenario_sheet_name = f'scenario_{20 if IS <= 20 else 40}'
    else:
        scenario_sheet_name = f'scenario'
    
    scenario_data = pd.read_excel(filename, scenario_sheet_name)
    distance_data = pd.read_excel(filename, 'distance', header=None)
    facility_cost_data = pd.read_excel(filename, 'facility_cost', header=None)
    resource_cost_data = pd.read_excel(filename, 'resource_cost', header=None)

    # 只存储所需的数据
    r.set('CF', json.dumps(facility_cost_data.loc[1, 1:3].tolist()))
    r.set('U', json.dumps(facility_cost_data.loc[2, 1:3].tolist()))
    r.set('H', json.dumps(distance_data.loc[1:IS+1, 1:IS+1].values.tolist()))
    r.set('V', json.dumps(resource_cost_data.loc[1, 1:3].tolist()))
    r.set('CP', json.dumps(resource_cost_data.loc[2, 1:3].tolist()))
    r.set('CT', json.dumps(resource_cost_data.loc[3, 1:3].tolist()))
    r.set('CH', json.dumps(resource_cost_data.loc[4, 1:3].tolist()))
    r.set('PU', json.dumps(resource_cost_data.loc[5, 1:3].tolist()))
    r.set('demand_raw', json.dumps(scenario_data.iloc[:, 1:IS + 1].values.tolist()))

    print("Data stored in Redis successfully!")

# Redis存储接口3
# def store_data_to_redis(filename, IS, Raw_data_flag):
#     # 连接到Redis
#     r = redis.Redis(host='localhost', port=6379, db=0)

#     # 读取Excel数据
#     if Raw_data_flag:
#         scenario_sheet_name = f'scenario_{20 if IS <= 20 else 40}'
#     else:
#         scenario_sheet_name = f'scenario'
    
#     scenario_data = pd.read_excel(filename, scenario_sheet_name)
#     distance_data = pd.read_excel(filename, 'distance', header=None)
#     facility_cost_data = pd.read_excel(filename, 'facility_cost', header=None)
#     resource_cost_data = pd.read_excel(filename, 'resource_cost', header=None)

#     # 将数据存储到Redis哈希中
#     cost_data = {
#         'CF': pickle.dumps(facility_cost_data.loc[1, 1:3].tolist()),
#         'U': pickle.dumps(facility_cost_data.loc[2, 1:3].tolist()),
#         'V': pickle.dumps(resource_cost_data.loc[1, 1:3].tolist()),
#         'CP': pickle.dumps(resource_cost_data.loc[2, 1:3].tolist()),
#         'CT': pickle.dumps(resource_cost_data.loc[3, 1:3].tolist()),
#         'CH': pickle.dumps(resource_cost_data.loc[4, 1:3].tolist()),
#         'PU': pickle.dumps(resource_cost_data.loc[5, 1:3].tolist())
#     }
#     for key, value in cost_data.items():
#         r.hset('cost_data', key, value)

#     # 将其他数据存储到Redis中
#     r.set('H', pickle.dumps(distance_data.loc[1:IS+1, 1:IS+1].values.tolist()))
#     r.set('demand_raw', pickle.dumps(scenario_data.iloc[:, 1:IS + 1].values.tolist()))

#     print("Data stored in Redis successfully!")


"""
从Redis数据库读取数据并进行处理，与之前的store_data_to_redis一一对应

本函数连接到Redis数据库，获取之前存储的各项数据，并根据需要进行处理。处理后的数据将用于随后的计算和分析。

:param IS: 场景数量，用于确定读取和处理数据的规模。
:param NS: 需要读取的数据点数量，用于按需读取数据。
:param AS: 物资种类的数量。
:param Food: 食品需求的比例因子。
:param Medicine: 药品需求的比例因子。
:return: 返回处理后的参数和数据数组，包括固定成本、存储容量、距离矩阵等。
"""
# Redis读取接口1
# def read_data_from_redis(IS, NS, AS, Food, Medicine):
#     # 连接到Redis
#     r = redis.Redis(host='localhost', port=6379, db=0, decode_responses=True)

#     # 一次性从Redis获取所有必要的数据
#     keys = ['scenario_data', 'distance_data', 'facility_cost_data', 'resource_cost_data']
#     scenario_data_json, distance_data_json, facility_cost_data_json, resource_cost_data_json = r.mget(keys)
    
#     # 直接读取JSON数据
#     scenario_data = pd.read_json(scenario_data_json)
#     distance_data = pd.read_json(distance_data_json)
#     facility_cost_data = pd.read_json(facility_cost_data_json)
#     resource_cost_data = pd.read_json(resource_cost_data_json)
    
#     # # 从Redis读取数据
#     # scenario_data_json = r.get('scenario_data')
#     # distance_data_json = r.get('distance_data')
#     # facility_cost_data_json = r.get('facility_cost_data')
#     # resource_cost_data_json = r.get('resource_cost_data')

#     # # 使用StringIO读取JSON数据
#     # scenario_data = pd.read_json(StringIO(scenario_data_json))
#     # distance_data = pd.read_json(StringIO(distance_data_json))
#     # facility_cost_data = pd.read_json(StringIO(facility_cost_data_json))
#     # resource_cost_data = pd.read_json(StringIO(resource_cost_data_json))

#     # 数据处理
#     demand_index = 0.03
#     pr = np.full(NS, 1/NS)

#     demand_raw = scenario_data.iloc[:NS, 1:IS + 1].to_numpy()
#     demand = np.rint(demand_raw * demand_index)
#     Dr = np.empty((NS, IS, AS))
#     Dr[:, :, 0] = demand_raw
#     Dr[:, :, 1] = np.rint(demand_raw * Food)
#     Dr[:, :, 2] = np.rint(demand_raw * Medicine)

#     CF = facility_cost_data.loc[1, 1:3].to_numpy()
#     U = facility_cost_data.loc[2, 1:3].to_numpy()
#     H = distance_data.loc[1:IS, 1:].to_numpy()
#     V = resource_cost_data.loc[1, 1:3].to_numpy()
#     CP = resource_cost_data.loc[2, 1:3].to_numpy()
#     CT = resource_cost_data.loc[3, 1:3].to_numpy()
#     CH = resource_cost_data.loc[4, 1:3].to_numpy()
#     PU = resource_cost_data.loc[5, 1:3].to_numpy()

#     D = np.rint(Dr * demand_index).transpose(0, 2, 1)

#     # print(CF, U, H, V, CP, CH, PU, CT, D, pr, demand)

#     return CF, U, H, V, CP, CH, PU, CT, D, pr, demand

# Redis读取接口2
def read_data_from_redis(IS, NS, AS, Food, Medicine):
    # 连接到Redis
    r = redis.Redis(host='localhost', port=6379, db=0, decode_responses=True)

    # 从Redis读取数据
    CF = json.loads(r.get('CF'))
    U = json.loads(r.get('U'))
    H = json.loads(r.get('H'))
    V = json.loads(r.get('V'))
    CP = json.loads(r.get('CP'))
    CT = json.loads(r.get('CT'))
    CH = json.loads(r.get('CH'))
    PU = json.loads(r.get('PU'))
    demand_raw = json.loads(r.get('demand_raw'))

    # 数据处理
    demand_index = 0.03
    pr = np.full(NS, 1/NS)

    demand_raw = np.array(demand_raw)[:NS, :IS]
    demand = np.rint(demand_raw * demand_index)
    
    Dr = np.empty((NS, IS, AS))
    Dr[:, :, 0] = demand_raw
    Dr[:, :, 1] = np.rint(demand_raw * Food)
    Dr[:, :, 2] = np.rint(demand_raw * Medicine)

    D = np.rint(Dr * demand_index).transpose(0, 2, 1)

    return CF, U, np.array(H)[:IS, :IS], V, CP, CH, PU, CT, D, pr, demand

# Redis读取接口3
# def read_data_from_redis(IS, NS, AS, Food, Medicine):
#     # 连接到Redis
#     r = redis.Redis(host='localhost', port=6379, db=0)

#     # 使用管道读取数据
#     with r.pipeline() as pipe:
#         for key in ['CF', 'U', 'V', 'CP', 'CT', 'CH', 'PU']:
#             pipe.hget('cost_data', key)
#         pipe.get('H')
#         pipe.get('demand_raw')
#         result = pipe.execute()

#     # 反序列化数据
#     cost_data_values = result[:7]
#     H = pickle.loads(result[7])
#     demand_raw = pickle.loads(result[8])

#     CF, U, V, CP, CT, CH, PU = [pickle.loads(value) for value in cost_data_values]

#     # 数据处理
#     demand_index = 0.03
#     pr = np.full(NS, 1/NS)

#     demand_raw = np.array(demand_raw)[:NS, :IS]
#     demand = np.rint(demand_raw * demand_index)
    
#     Dr = np.empty((NS, IS, AS))
#     Dr[:, :, 0] = demand_raw
#     Dr[:, :, 1] = np.rint(demand_raw * Food)
#     Dr[:, :, 2] = np.rint(demand_raw * Medicine)

#     D = np.rint(Dr * demand_index).transpose(0, 2, 1)

#     return CF, U, np.array(H)[:IS, :IS], V, CP, CH, PU, CT, D, pr, demand

# Pandas读取数据接口
def read_data(filename, IS, NS, AS, Food, Medicine, Raw_data_flag):
    """
    从Excel文件直接用pandas读取数据并返回参数和数据数组。注意与之前的Redis可以切换着使用

    本函数直接从Excel文件中读取所有必要的数据，包括场景数据、距离数据、设施成本数据和资源成本数据。读取的数据将用于构建和求解应急物资配置问题的优化模型。

    :param filename: Excel文件的路径和名称。
    :param IS: 场景数量，用于确定读取Excel中的哪个场景数据。
    :param NS: 需要读取的数据点数量，用于按需读取数据。
    :param AS: 物资种类的数量。
    :param Food: 食品需求的比例因子。
    :param Medicine: 药品需求的比例因子。
    :param Raw_data_flag: 原始数据标志，如果为True，则读取原始格式的数据；否则读取常规格式的数据。
    :return: 返回处理后的参数和数据数组，包括固定成本、存储容量、距离矩阵等。
    """
    # 这个部分的数据处理主要用于保证示例数据计算结果与论文所示相同，随机生成的算例不受影响
    demand_index = 0.03
    if Raw_data_flag:
        if IS <= 20:
            scenario_sheet_name = f'scenario_20'
        else:
            scenario_sheet_name = f'scenario_40'
    else:
        scenario_sheet_name = f'scenario'
        
    distance_sheet_name = f'distance'
    facility_cost_sheet_name = f'facility_cost'
    resource_cost_sheet_name = f'resource_cost'
    
    scenario_data = pd.read_excel(filename, scenario_sheet_name)
    distance_data = pd.read_excel(filename, distance_sheet_name, header=None)
    facility_cost_data = pd.read_excel(filename, facility_cost_sheet_name, header=None)
    resource_cost_data = pd.read_excel(filename, resource_cost_sheet_name, header=None)

    # probability
    pr = np.full(NS, 1/NS)

    # demand
    demand_raw = scenario_data.iloc[:NS, 1:IS + 1].to_numpy()
    demand = np.rint(demand_raw * demand_index)
    # print(demand)
    Dr = np.empty((NS, IS, AS))
    Dr[:, :, 0] = demand_raw
    Dr[:, :, 1] = np.rint((demand_raw * Food))
    Dr[:, :, 2] = np.rint((demand_raw * Medicine))

    # Fixed cost, Storage capacity, volume of items
    CF = facility_cost_data.loc[1, 1:3].to_numpy()
    U = facility_cost_data.loc[2, 1:3].to_numpy()

    # Distance
    H = distance_data.loc[1:IS, 1:].to_numpy()

    # volume of items
    V = resource_cost_data.loc[1, 1:3].to_numpy()
    # Unit procurement price
    CP = resource_cost_data.loc[2, 1:3].to_numpy()
    # Unit transportation cost
    CT = resource_cost_data.loc[3, 1:3].to_numpy()
    # Unit holding cost
    CH = resource_cost_data.loc[4, 1:3].to_numpy()
    # Unit penalty cost
    PU = resource_cost_data.loc[5, 1:3].to_numpy()

    D = np.rint(Dr * demand_index).transpose(0, 2, 1)
    
    
    # print(CF, U, PU, V, CP, CH, G, CT, D, pr, demand)

    return CF, U, H, V, CP, CH, PU, CT, D, pr, demand

# 测试数据读取接口
def read_data_old(filename, IS, NS, AS, Food, Medicine):
    """
    测试接口，将在后续更新修正
    从Excel文件读取旧格式的数据并返回参数和数据数组。

    该函数用于处理遗留的Excel数据格式，提取场景数据、需求数据以及其他相关成本数据。这些数据是应急物资配置问题求解所必需的。

    :param filename: Excel文件的路径和名称。
    :param IS: 场景数量，用于确定读取Excel中的哪个场景数据。
    :param NS: 需要读取的数据点数量，用于按需读取数据。
    :param AS: 物资种类的数量。
    :param Food: 食品需求的比例因子。
    :param Medicine: 药品需求的比例因子。
    :return: 返回处理后的参数和数据数组，包括固定成本、存储容量、距离矩阵等。
    """

    scenario_sheet_name = f'scenario_{IS}'
    probability_column_name = f'w{NS // 100}'

    scenario_data = pd.read_excel(filename, scenario_sheet_name)
    main_data = pd.read_excel(filename, 'main_data', header=None)

    # probability
    pr = scenario_data[probability_column_name].to_numpy()

    # demand
    demand_index = scenario_data.iloc[:NS, 1:IS + 1].to_numpy()
    demand = np.rint(demand_index * 0.03)
    Dr = np.empty((NS, IS, AS))
    Dr[:, :, 0] = demand_index
    Dr[:, :, 1] = np.rint((Dr[:, :, 0] * Food))
    Dr[:, :, 2] = np.rint((Dr[:, :, 0] * Medicine))

    # Fixed cost, Storage capacity, volume of items
    CF = main_data.loc[1, 1:3].to_numpy()
    U = main_data.loc[2, 1:3].to_numpy()

    # Distance
    PU = main_data.loc[8:(7 + IS), 1:].to_numpy()

    # volume of items
    V = main_data.loc[1, 6:8].to_numpy()
    # Unit procurement price
    CP = main_data.loc[2, 6:8].to_numpy()
    # Unit transportation cost
    CT = main_data.loc[3, 6:8].to_numpy()
    # Unit holding cost
    CH = main_data.loc[4, 6:8].to_numpy()
    # Unit penalty cost
    PU = main_data.loc[5, 6:8].to_numpy()
    
    D = np.rint(Dr * 0.03).transpose(0, 2, 1)

    return CF, U, PU, V, CP, CH, PU, CT, D, pr, demand

def generate_script_name(data_preprocess_methods, cluster_methods, sample_methods):
    """
    生成用于聚类图表文件名的脚本名称。

    根据数据预处理方法、聚类方法和样本生成方法的组合，生成一个独特的脚本名称。这个名称将用于保存和识别聚类图表文件。

    :param data_preprocess_methods: 使用的数据预处理方法。
    :param cluster_methods: 使用的聚类方法。
    :param sample_methods: 使用的样本生成方法。
    :return: 返回生成的脚本名称字符串。
    """
    if data_preprocess_methods:
        script_name = f"{data_preprocess_methods}_{cluster_methods}_{sample_methods}"
    else:
        script_name = f"{cluster_methods}_{sample_methods}"
    return script_name

def append_df_to_excel(filename, df, sheet_name='result', startrow=None ,index=False ,header=False , **to_excel_kwargs):
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    """
    将DataFrame [df] 追加到已存在的Excel文件 [filename] 的 [sheet_name] 工作表中。
    如果 [filename] 不存在，此函数将创建它。

    参数:
      filename : 文件路径或已存在的Excel文件
                 (示例: '/path/to/file.xlsx')
      df : 需要保存到工作簿的dataframe
      sheet_name : 将包含DataFrame数据的工作表名称。
                   (默认: 'Sheet1')
      startrow : 左上角单元格行开始存放数据框架的位置。
                 默认情况下 (startrow=None) 计算现有DF的最后一行
                 并写入到下一行...
      to_excel_kwargs : 将传递给 `DataFrame.to_excel()` 的参数
                        [可以是一个字典]

    返回: 无

    示例:
    append_df_to_excel('output.xlsx', df)
    """

    # Load existing workbook or create a new one
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(filename)

    # Check if sheet exists, if not create it
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    # Get the active sheet
    sheet = wb[sheet_name]

    # Calculate starting row
    if startrow is None:
        startrow = sheet.max_row if sheet.max_row > 0 else 1

    # Convert the DataFrame to rows
    # rows = list(dataframe_to_rows(df, header=True))
    rows = list(dataframe_to_rows(df, index=index, header=header))

    # Write DataFrame rows to Excel sheet
    for row in rows:
        sheet.append(row)

    # Save the workbook
    wb.save(filename)

def plot_cluster_sampling(demand_transformed, cluster_labels, sample, save_directory, script_name, m):
    """
    绘制聚类采样结果图。

    参数:
    demand_transformed : 转换后的需求数据。
    cluster_labels : 聚类标签。
    sample : 采样点索引列表。
    save_directory : 图片保存目录。
    script_name : 脚本名称，用作保存图片的一部分文件名。
    m : 样本编号，用作保存图片的一部分文件名。
    
    功能:
    在二维平面上绘制聚类结果，并突出显示采样点。
    将绘制的图表保存到指定目录。
    """
    # 绘制聚类结果和分层采样的样本点
    plt.figure(figsize=(10, 8),dpi=300)
    for i, label in enumerate(np.unique(cluster_labels)):
        plt.scatter(
            demand_transformed[cluster_labels == label, 0],
            demand_transformed[cluster_labels == label, 1],
            label=f'Cluster {i}'
        )

    # 假设sample包含了所有采样点的索引
    for s in sample:
        plt.scatter(
            demand_transformed[s, 0],
            demand_transformed[s, 1],
            c='red', # 标记颜色
            edgecolor='k',
            marker='x', # 标记样式
            s=100 # 设置标记大小为100 point^
        )

    plt.title('Sampling of Clusters')
    plt.xlabel('Feature 1')
    plt.ylabel('Feature 2')
    plt.legend()

    # 检查保存目录是否存在，如果不存在则创建
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # 保存图片到指定目录
    plt.savefig(os.path.join(save_directory, f'{script_name}_sample_{m+1}.jpg'), format='jpg')
    plt.close()  # 关闭图形界面，防止内存泄露

def plot_cluster_3d_sampling(demand_transformed, cluster_labels, sample, save_directory, script_name, m):
    """
    绘制3D聚类采样结果图。

    参数:
    demand_transformed : 转换后的需求数据。
    cluster_labels : 聚类标签。
    sample : 采样点索引列表。
    save_directory : 图片保存目录。
    script_name : 脚本名称，用作保存图片的一部分文件名。
    m : 样本编号，用作保存图片的一部分文件名。

    功能:
    在三维空间中绘制聚类结果，并突出显示采样点。
    将绘制的图表保存到指定目录。
    """
    # 创建一个新的图和一个三维轴
    fig = plt.figure(figsize=(10, 8), dpi=300)
    ax = fig.add_subplot(111, projection='3d')

    # 绘制聚类结果的散点图
    for i, label in enumerate(np.unique(cluster_labels)):
        ax.scatter(
            demand_transformed[cluster_labels == label, 0],  # x轴坐标
            demand_transformed[cluster_labels == label, 1],  # y轴坐标
            demand_transformed[cluster_labels == label, 2],  # z轴坐标
            label=f'Cluster {i}'
        )

    # 假设sample包含了所有采样点的索引
    for s in sample:
        ax.scatter(
            demand_transformed[s, 0],  # x轴坐标
            demand_transformed[s, 1],  # y轴坐标
            demand_transformed[s, 2],  # z轴坐标
            c='red',  # 标记颜色
            facecolor='k',
            marker='x',  # 标记样式
            s=100  # 设置标记大小为100 point^
        )

    # 设置图表标题和坐标轴标签
    ax.set_title('3D Sampling of Clusters')
    ax.set_xlabel('Feature 1')
    ax.set_ylabel('Feature 2')
    ax.set_zlabel('Feature 3')

    # 设置视角
    ax.view_init(elev=20., azim=45)  # 设置斜视图

    # 显示图例
    ax.legend()

    # 检查保存目录是否存在，如果不存在则创建
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # 保存图片到指定目录
    # save_path = os.path.join(save_directory, f'{script_name}_sample_3D_{m+1}.jpg')
    # plt.savefig(save_path, format='jpg')
    plt.savefig(os.path.join(save_directory, f'{script_name}_sample_3D_{m+1}.jpg'), format='jpg')
    plt.close()  # 关闭图形界面，防止内存泄露

def plot_cluster(demand_transformed, cluster_labels, save_directory, script_name):
    """
    绘制聚类结果图。

    参数:
    demand_transformed : 转换后的需求数据。
    cluster_labels : 聚类标签。
    save_directory : 图片保存目录。
    script_name : 脚本名称，用作保存图片的一部分文件名。

    功能:
    在二维平面上绘制聚类结果。
    将绘制的图表保存到指定目录。
    """
    plt.figure(figsize=(10, 8), dpi=300)
    # 绘制聚类结果
    for i, label in enumerate(np.unique(cluster_labels)):
        plt.scatter(
            demand_transformed[cluster_labels == label, 0],
            demand_transformed[cluster_labels == label, 1],
            label=f'Cluster {i}'
        )

    plt.title('Clustering Results')
    plt.xlabel('Feature 1')
    plt.ylabel('Feature 2')
    plt.legend()

    # 检查保存目录是否存在，如果不存在则创建
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # 保存图片到指定目录
    plt.savefig(os.path.join(save_directory, f'{script_name}_cluster.jpg'), format='jpg')
    plt.close()  # 关闭图形界面，防止内存泄露

def plot_cluster_3d(demand_transformed, cluster_labels, save_directory, script_name):
    """
    绘制3D聚类结果图。

    参数:
    demand_transformed : 转换后的需求数据。
    cluster_labels : 聚类标签。
    save_directory : 图片保存目录。
    script_name : 脚本名称，用作保存图片的一部分文件名。

    功能:
    在三维空间中绘制聚类结果。
    将绘制的图表保存到指定目录。
    """
    # 创建一个新的图和一个三维轴
    fig = plt.figure(figsize=(10, 8), dpi=300)
    ax = fig.add_subplot(111, projection='3d')

    # 绘制聚类结果的散点图
    for i, label in enumerate(np.unique(cluster_labels)):
        ax.scatter(
            demand_transformed[cluster_labels == label, 0],  # x轴坐标
            demand_transformed[cluster_labels == label, 1],  # y轴坐标
            demand_transformed[cluster_labels == label, 2],  # z轴坐标
            label=f'Cluster {i}'
        )

    # 设置图表标题和坐标轴标签
    ax.set_title('3D Clustering Results')
    ax.set_xlabel('Feature 1')
    ax.set_ylabel('Feature 2')
    ax.set_zlabel('Feature 3')

    # 设置视角
    ax.view_init(elev=20., azim=45)

    # 显示图例
    ax.legend()

    # 检查保存目录是否存在，如果不存在则创建
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # 保存图片到指定目录
    save_path = os.path.join(save_directory, f'{script_name}_cluster_3D.jpg')
    plt.savefig(save_path, format='jpg')
    plt.close()  # 关闭图形界面，防止内存泄露

def generate_cluster_plots(demand_transformed, cluster_labels, Graphs_cluster_save_directory, script_name, graph_type):
    """
    根据指定的图表类型生成聚类图表。

    参数:
    demand_transformed : 转换后的需求数据。
    cluster_labels : 聚类的标签。
    Graphs_cluster_save_directory : 保存聚类图表的目录。
    script_name : 脚本名称，用于图表的文件名。
    graph_type : 图表类型，'2d' 或 '3d'。

    功能:
    根据给定的需求数据和聚类标签，生成并保存聚类图表。
    根据图表类型，选择调用2D或3D绘图函数。
    """
    
    if graph_type == '2d':
        # 调用2D绘图函数
        plot_cluster(demand_transformed, cluster_labels, Graphs_cluster_save_directory, script_name)
    elif graph_type == '3d':
        # 调用3D绘图函数
        plot_cluster_3d(demand_transformed, cluster_labels, Graphs_cluster_save_directory, script_name)
    else:
        raise ValueError("Invalid graph_type: {}. Choose '2d' or '3d'.".format(graph_type))

def generate_sample_plots(demand_transformed, samples_info, cluster_labels, Graphs_sample_save_directory, graph_type):
    """ 
    生成每个样本的聚类图。根据指定的图形类型（2D或3D），对不同样本进行聚类并绘制图形。

    参数说明：

    demand_transformed: 转换后的需求数据。
    samples_info: 包含样本信息的列表，每个元素是一个包含样本、脚本名称和其他信息的元组。
    cluster_labels: 聚类标签。
    Graphs_sample_save_directory: 图形保存的目录路径。
    graph_type: 图形类型（'2d' 或 '3d'）。
    异常处理：
    如果图形类型不是 '2d' 或 '3d'，则抛出 ValueError。
    """ 
    # Generate cluster plots for each sample
    for sample, script_name, m in samples_info:
        if graph_type == '2d':
            plot_cluster_sampling(demand_transformed, cluster_labels, sample, Graphs_sample_save_directory, script_name, m)
        elif graph_type == '3d':
            plot_cluster_3d_sampling(demand_transformed, cluster_labels, sample, Graphs_sample_save_directory, script_name, m)
        else:
            raise ValueError("Invalid plot_type: {}. Choose '2d' or '3d'.".format(graph_type))

def calculate_gap(ff, MS, gurobi_opt):
    """ 
    计算和返回GAP百分比，用来评估模型的优化效果。

    参数说明：

    ff: 各样本解的数组。
    MS: 样本数量。
    gurobi_opt: 使用Gurobi优化得到的全局最优解。
    异常处理：

    如果样本数量 MS 小于等于0，则抛出 ValueError。
    如果全局最优解 gurobi_opt 小于等于0，则抛出 ValueError。
    返回值：
    GAP百分比。
    """ 
    # Check for valid inputs
    if MS <= 0:
        raise ValueError("Number of samples (MS) must be greater than 0.")
    if gurobi_opt <= 0:
        raise ValueError("Global best result (gurobi_opt) must be greater than 0.")
    
    # Calculate average of the solutions
    ave_f = np.sum(ff) / MS

    # Calculate GAP
    gap_percentage = (gurobi_opt - ave_f) / gurobi_opt * 100  # GAP as a percentage

    return gap_percentage

def save_and_print_results(script_name, Output_file, Vx, Vy, IS, NS, MS, SS_SAA, opt_f, elapsed_time, cluster_num = 0, gap = 0):
    """
    保存结果到Excel文件并打印。该函数还将位置和库存数据分别保存到Excel的特定工作表中。

    参数说明：

    script_name: 脚本名称。
    Output_file: 输出文件路径。
    Vx: 位置数据。
    Vy: 库存数据。
    IS: 输入大小。
    NS: 节点大小。
    MS: 模型大小。
    SS_SAA: SAA方法的样本大小。
    opt_f: 优化结果。
    elapsed_time: 耗时。
    cluster_num (可选): 聚类数量。
    gap (可选): 计算得到的GAP值。
    详细功能：

    为不同的输入和节点大小创建独立的工作表。
    将基础结果数据和更详细的数据（如位置和库存信息）分别保存到指定的工作表中。
    """
    # Naming the sheets with IS and NS values
    results_sheet_name = f'results_IS_{IS}_NS_{NS}'
    details_sheet_name = f'details_IS_{IS}_NS_{NS}'

    # 创建一个DataFrame来组织需要输出的数据
    result_df = pd.DataFrame([[script_name, IS, NS, MS, SS_SAA, float(opt_f), elapsed_time, cluster_num, gap]])
    detail_df = pd.DataFrame([[script_name, IS, NS, MS, SS_SAA, float(opt_f), elapsed_time, gap]])

    # 将数据输出到Excel的特定列，只有一列
    append_df_to_excel(Output_file, result_df, sheet_name=results_sheet_name, index=False, header=False, startrow=0)

    location_df = pd.DataFrame(Vx)
    inventory_df = pd.DataFrame(Vy).T
    
    start_row = 1

    # Save costs data
    append_df_to_excel(Output_file, detail_df, sheet_name=details_sheet_name, index=False, header=False, startrow=0)

    # Calculate next start column for location data
    location_startcol = detail_df.shape[1] + 2  # Assuming 2 column gap for readability
    append_df_to_excel(Output_file, location_df, sheet_name=details_sheet_name, index=True, header=False, startrow=start_row, startcol=location_startcol)

    # Calculate next start column for inventory data
    inventory_startcol = location_startcol + location_df.shape[1] + 2  # Assuming 2 column gap for readability
    append_df_to_excel(Output_file, inventory_df, sheet_name=details_sheet_name, index=True, header=False, startrow=start_row, startcol=inventory_startcol)

    # # Calculate start row for elapsed time data
    # elapsed_time_startrow = start_row + max(detail_df.shape[0], inventory_df.shape[0], location_df.shape[0]) + 2  # Assuming 2 row gap for readability
    # append_df_to_excel(Output_file, elapsed_time_df, sheet_name='details', index=True, header=True, startrow=elapsed_time_startrow)